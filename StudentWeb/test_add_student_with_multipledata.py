import requests
import json
import jsonpath
import openpyxl



def test_add_multiple_students():
    url="https://thetestingworldapi.com/api/studentsDetails"
    f=open("C:/scriptExample/Pytest/StudentWeb/student.json","r")

    #reading info from xls
    workBook=openpyxl.load_workbook("C:/scriptExample/Pytest/StudentWeb/multiple_student.xlsx")
    sheet=workBook['Sheet1']
    rows=sheet.max_row  #how many rows in the .xls

    student_data = json.loads((f.read()))
    print(f" \n student info is {student_data} \n ")

    for i in range(2,rows+1):  #to read from 2nd row to last line
        cells_first_name=sheet.cell(row=i,column=1)
        cells_middle_name = sheet.cell(row=i, column=2)
        cells_last_name = sheet.cell(row=i, column=3)
        cells_date_of_birth = sheet.cell(row=i, column=4)

        student_data['first_name']=cells_first_name.value
        student_data['middle_name'] = cells_middle_name.value
        student_data['last_name'] = cells_last_name.value
        student_data['date_of_birth'] = cells_date_of_birth.value

        response = requests.post(url, student_data)
        print(f"Response status is {response.status_code}")
        print(f" \t Response text is {response.text}")
        assert response.status_code == 201





